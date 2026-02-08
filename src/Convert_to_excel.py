import openpyxl
from openpyxl.drawing.image import Image
# import os
from datetime import datetime

wb = openpyxl.load_workbook("data/APPLICATION-for-MARRIAGE-LICENSE.xlsx")
sheet = wb.active

# ------VARIABLES FOR INPUT------------
# pwede kahit isang parent lang
mother_female_first_name = "Maria"
mother_female_last_name = "Osawa"
father_female_first_name = "Johnny"
father_female_last_name = "Sins"
parent_female_address = "Solano, Nueva Vizcaya"

mother_male_first_name = "Siobe"
mother_male_last_name = "Nanonaha"
father_male_first_name = "Alnico"
father_male_last_name = "Fernandez"

parent_classification_mother = "Mother"
parent_classification_father = "Father"

female_first_name = "Shane"
female_last_name = "Mendoza"
female_birth_place = "Bayombong, Nueva Vizcaya"
female_address = "Solano, Nueva Vizcaya"

male_first_name = "Jefferson"
male_last_name = "Aliggoy"
male_birth_place = "Diadi, Nueva Vizcaya"
male_residence = "Solano, Nueva Vizcaya"
male_age = 23
female_age = 23

date_day = "8"
date_month = "February"
date_year = "####"

employee_name = "Mark John Barbieto"

year = 2016
number = 17

# Conversion function: cm → pixels


def cm_to_pixels(cm):
    return int((cm / 2.54) * 96)


couple_img = Image("data/couple_img.png")
couple_img.height = cm_to_pixels(3.75)
couple_img.width = cm_to_pixels(5.73)
# Select the "Notice" sheet (even if it's not active)
notice_sheet = wb["Notice"]
notice_sheet.add_image(couple_img, "T11")

wb["APPLICATION"]["B22"] = father_male_first_name
wb["APPLICATION"]["L22"] = father_male_last_name
wb["APPLICATION"]["B26"] = mother_male_first_name
wb["APPLICATION"]["K26"] = mother_male_last_name

wb["APPLICATION"]["U22"] = father_female_first_name
wb["APPLICATION"]["AC22"] = father_female_last_name
wb["APPLICATION"]["U26"] = mother_female_first_name
wb["APPLICATION"]["AD26"] = mother_female_last_name


wb["APPLICATION"]["U30"] = mother_female_first_name
wb["APPLICATION"]["AD30"] = mother_female_last_name


wb["APPLICATION"]["U34"] = parent_female_address
wb["APPLICATION"]["U31"] = parent_classification_mother


wb["APPLICATION"]["U8"] = female_first_name
wb["APPLICATION"]["U10"] = female_last_name
wb["APPLICATION"]["U12"] = female_birth_place
wb["APPLICATION"]["U15"] = female_address


wb["APPLICATION"]["B8"] = male_first_name
wb["APPLICATION"]["B10"] = male_last_name
wb["APPLICATION"]["B12"] = male_birth_place
wb["APPLICATION"]["B15"] = male_residence

wb["APPLICATION"]["N11"] = male_age
wb["APPLICATION"]["AF11"] = female_age

wb["APPLICATION"]["B37"] = date_day
wb["APPLICATION"]["E37"] = date_month
wb["APPLICATION"]["L37"] = date_year

wb["APPLICATION"]["X3"] = year
wb["APPLICATION"]["AD3"] = number

wb["APPLICATION"]["F4"] = employee_name


sheet_name = None

if (18 <= female_age <= 20 and male_age >= 25):
    sheet_name = "CONSENT F"
elif (18 <= male_age <= 20 and female_age >= 25):
    sheet_name = "CONSENT M"
elif (18 <= female_age <= 20 and 18 <= male_age <= 20):
    sheet_name = "CONSENT M&F"
elif (21 <= female_age <= 24 and male_age >= 25):
    sheet_name = "ADVICE F"
elif (21 <= male_age <= 24 and female_age >= 25):
    sheet_name = "ADVICE M"
elif (21 <= female_age <= 24 and 21 <= male_age <= 24):
    sheet_name = "ADVICE M&F"
elif (21 <= male_age <= 24 and 18 <= female_age <= 20):
    sheet_name = "ADVICE M-CONSENT F"
elif (21 <= female_age <= 24 and 18 <= male_age <= 20):
    sheet_name = "ADVICE F-CONSENT M"

# --- After your existing sheet_name logic ---

if sheet_name:
    if female_address == "Solano, Nueva Vizcaya" and male_residence == "Solano, Nueva Vizcaya":
        for sheet in wb.sheetnames:
            if sheet not in (sheet_name, "APPLICATION", "Notice"):
                del wb[sheet]
    else:
        for sheet in wb.sheetnames:
            if sheet not in (sheet_name, "APPLICATION", "Notice", "AddressBACKnotice", "EnvelopeAddress"):
                del wb[sheet]

    # Save with timestamped filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"Excel/{sheet_name}_{timestamp}.xlsx"
    wb.save(output_filename)
